
import os
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

# 📂 Definição dos arquivos (localizados na raiz do projeto)
arquivo_origem = "alunos_turmas.xlsx"  # Arquivo de origem
arquivo_destino = "atas_assinaturas.xlsx"  # Arquivo de destino

# 📅 Obtém a data atual formatada
data_atual = datetime.now().strftime("%d/%m/%Y")

# 🔁 Mapeamento de abas (ajuste conforme necessário)
mapa_abas = {
    "EM45-1A-I": "1 ano A",
    "EM45-1B-I": "1 ano B"
}

# 🚀 Verifica se os arquivos existem
if os.path.exists(arquivo_origem) and os.path.exists(arquivo_destino):
    print("Os arquivos foram encontrados!")

    # 📖 Carrega os arquivos
    df_origem = pd.read_excel(arquivo_origem, sheet_name=None)  # Lê todas as abas
    wb_destino = load_workbook(arquivo_destino)

    for origem, destino in mapa_abas.items():  # Para cada aba no mapeamento
        if origem in df_origem.keys() and destino in wb_destino.sheetnames:
            ws_destino = wb_destino[destino]

            # 🧹 Limpa os dados antigos a partir da linha 3
            for row in ws_destino.iter_rows(min_row=3, max_row=ws_destino.max_row, min_col=1, max_col=ws_destino.max_column):
                for cell in row:
                    cell.value = None

            # 🔄 Copia os dados da origem para o destino
            df = df_origem[origem]
            for i, row in enumerate(df.values, start=3):  # Começa na linha 3
                for j, value in enumerate(row, start=1):
                    ws_destino.cell(row=i, column=j, value=value)

            # 📝 Atualiza o título e a data
            ws_destino["A2"] = f"Lista de Alunos - {destino}"
            ws_destino["E2"] = f"Atualizado em: {data_atual}"

    # 💾 Salva as alterações
    wb_destino.save(arquivo_destino)
    print(f"✅ Arquivo {arquivo_destino} atualizado com sucesso!")
else:
    print("❌ Um ou ambos os arquivos não foram encontrados. Verifique os nomes e tente novamente.")
